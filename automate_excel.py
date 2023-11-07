from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Security.xlsx')
ws = wb['SCA']
# Access the value
#print(ws['A1'].value)
# Change the value
ws['A2'].value = "Nishank"
#print(wb.sheetnames)
# Refer to another sheet
# Need to save the workbook
wb.save('Security.xlsx')
ws = wb['SAST']
#print(ws['A2'].value)
wb.create_sheet(title='Test Worksheet', index=1)
for row in range(1,5):
    for col in range (1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
# merge cells
ws = wb['SCA']
ws.merge_cells("A3:B3")
# Need to save the workbook
wb.save('Security.xlsx')